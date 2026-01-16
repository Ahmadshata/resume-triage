import argparse
import csv
import datetime as dt
import re
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

import fitz  # PyMuPDF

# Excel formatting
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# EASY CONFIG (edit these)
# =============================================================================

# Minimum DevOps experience required (years)
MIN_DEVOPS_YEARS: float = 3.0

# Required keywords that MUST appear in EXPERIENCE entries (case-insensitive)
# Example: {"Kubernetes", "AWS"} or {"Kubernetes", "AWS", "Terraform"}
REQUIRED_EXPERIENCE_KEYWORDS: Set[str] = {"Kubernetes", "AWS", "Helm"}

# How many lines around a match to include in a snippet (1 means +/-1 line)
SNIPPET_CONTEXT_LINES: int = 1


# =============================================================================
# Other config
# =============================================================================

EXPERIENCE_HEADINGS = {
    "experience",
    "work experience",
    "professional experience",
    "employment history",
    "career history",
}

STOP_HEADINGS = {
    "skills",
    "technical skills",
    "education",
    "projects",
    "certifications",
    "certification",
    "training",
    "summary",
    "profile",
    "publications",
    "courses",
    "languages",
    "volunteering",
    "interests",
}

# Use regex with word boundaries to avoid false positives like "certification" containing "iti"
EXCLUDE_PATTERNS = [
    re.compile(r"\biti\b", re.IGNORECASE),
    re.compile(r"\bnti\b", re.IGNORECASE),
    re.compile(r"\bsprints\b", re.IGNORECASE),
    re.compile(r"\bdepi\b", re.IGNORECASE),
    re.compile(r"information\s+technology\s+institute", re.IGNORECASE),
    re.compile(r"national\s+technology\s+institute", re.IGNORECASE),
]

DEVOPS_KEYWORDS = {
    "devops",
    "sre",
    "site reliability",
    "platform engineer",
    "platform engineering",
    "infrastructure",
    "cloud engineer",
    "cloud engineering",
    "kubernetes",
    "terraform",
    "terragrunt",
    "ci/cd",
    "cicd",
    "jenkins",
    "github actions",
    "helm",
    "eks",
    "docker",
    "ansible",
    "prometheus",
    "grafana",
    "argo",
    "argo cd",
    "gitops",
    "linux",
    "iac",
    "infrastructure as code",
    "cloudformation",
}

EDUCATION_HINTS = {"bachelor", "master", "masters", "degree", "faculty", "university", "education"}

JOB_TITLE_HINTS = {
    "engineer",
    "developer",
    "administrator",
    "architect",
    "consultant",
    "specialist",
    "lead",
    "manager",
    "intern",
    "head",
}

MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}

DATE_RANGE_PATTERN = re.compile(
    r"(?P<start>(?:[A-Za-z]{3,9}\s+\d{4})|(?:\d{1,2}[/-]\d{4})|(?:\d{4}))\s*"
    r"(?:-|–|—|to)\s*"
    r"(?P<end>(?:[A-Za-z]{3,9}\s+\d{4})|(?:\d{1,2}[/-]\d{4})|(?:\d{4})|"
    r"(?:present|current|now))",
    re.IGNORECASE,
)


# -----------------------------
# Data models
# -----------------------------

@dataclass
class Entry:
    lines: List[Tuple[int, str]]  # (page_number, line)

    def text(self) -> str:
        return "\n".join(line for _, line in self.lines).strip()

    def head(self, n: int = 3) -> str:
        out: List[str] = []
        for _, line in self.lines:
            s = line.strip()
            if s:
                out.append(s)
            if len(out) >= n:
                break
        return " | ".join(out)


@dataclass
class Role:
    title: str
    start: dt.date
    end: dt.date
    months_added: int


# -----------------------------
# Text extraction
# -----------------------------

def extract_text_by_page(pdf_path: Path) -> Tuple[List[str], bool]:
    doc = fitz.open(pdf_path)
    pages = [page.get_text("text") for page in doc]
    used_ocr = False

    # Optional OCR fallback if text extraction looks empty/scanned
    text = "\n".join(pages).strip()
    if len(text) < 500:
        ocr = try_ocr(doc)
        if ocr:
            pages = ocr
            used_ocr = True

    return pages, used_ocr


def try_ocr(doc: fitz.Document) -> Optional[List[str]]:
    try:
        import pytesseract  # type: ignore
        from PIL import Image  # type: ignore
    except Exception:
        return None

    ocr_pages: List[str] = []
    for page in doc:
        pix = page.get_pixmap(dpi=200)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        ocr_pages.append(pytesseract.image_to_string(img))
    return ocr_pages


def iter_lines_with_pages(pages: Sequence[str]) -> Iterable[Tuple[int, str]]:
    for i, page_text in enumerate(pages, start=1):
        for line in page_text.splitlines():
            yield i, line.rstrip()


def normalize_heading(line: str) -> str:
    return re.sub(r"[^a-z\s]", "", line.lower()).strip()


# -----------------------------
# Experience extraction
# -----------------------------

def capture_experience_by_heading(pages: Sequence[str], max_back_lines: int = 200) -> List[Tuple[int, str]]:
    """
    Heading-based capture, with a backward window to handle PDFs where extracted text order is odd
    (experience content appears before the "WORK EXPERIENCE" line).
    """
    lines = list(iter_lines_with_pages(pages))
    if not lines:
        return []

    exp_idxs: List[int] = []
    for idx, (_, line) in enumerate(lines):
        if normalize_heading(line) in EXPERIENCE_HEADINGS:
            exp_idxs.append(idx)

    if not exp_idxs:
        return []

    captured: List[Tuple[int, str]] = []

    for h_idx in exp_idxs:
        # Backward window
        start_back = max(0, h_idx - max_back_lines)
        for i in range(start_back, h_idx):
            _, l = lines[i]
            if normalize_heading(l) in STOP_HEADINGS:
                start_back = i + 1
        captured.extend(lines[start_back:h_idx])

        # Forward capture
        i = h_idx + 1
        while i < len(lines):
            _, l = lines[i]
            if normalize_heading(l) in STOP_HEADINGS:
                break
            captured.append(lines[i])
            i += 1

    # De-duplicate while preserving order
    seen: Set[Tuple[int, str]] = set()
    out: List[Tuple[int, str]] = []
    for item in captured:
        if item not in seen:
            seen.add(item)
            out.append(item)
    return out


def split_entries_from_lines(experience_lines: List[Tuple[int, str]]) -> List[Entry]:
    """
    Split into entries using blank lines as separators.
    """
    entries: List[Entry] = []
    current: List[Tuple[int, str]] = []

    for page_num, line in experience_lines:
        if not line.strip():
            if current:
                entries.append(Entry(lines=current))
                current = []
            continue
        current.append((page_num, line))

    if current:
        entries.append(Entry(lines=current))

    return [e for e in entries if e.text()]


def is_date_range_line(line: str) -> bool:
    return DATE_RANGE_PATTERN.search(line) is not None


def build_date_based_entries(pages: Sequence[str]) -> List[Entry]:
    """
    Create entries that start at date ranges (e.g., "Feb 2024 - Present")
    and continue until the next date range or a stop heading.
    """
    lines = list(iter_lines_with_pages(pages))
    entries: List[Entry] = []
    current: List[Tuple[int, str]] = []
    capturing = False

    for page_num, line in lines:
        if is_date_range_line(line):
            if current:
                entries.append(Entry(lines=current))
            current = [(page_num, line)]
            capturing = True
            continue

        if not capturing:
            continue

        if normalize_heading(line) in STOP_HEADINGS:
            if current:
                entries.append(Entry(lines=current))
            current = []
            capturing = False
            continue

        current.append((page_num, line))

    if current:
        entries.append(Entry(lines=current))

    return [e for e in entries if e.text()]


def is_excluded(entry_text: str) -> bool:
    return any(p.search(entry_text) for p in EXCLUDE_PATTERNS)


def is_devops_related(entry_text: str) -> bool:
    lower = entry_text.lower()
    return any(k in lower for k in DEVOPS_KEYWORDS)


def is_experience_entry(entry: Entry) -> bool:
    """
    Conservative filter to avoid counting education/certificates as experience.
    """
    text = entry.text().lower()

    if not DATE_RANGE_PATTERN.search(entry.text()):
        return False

    if any(h in text for h in EDUCATION_HINTS):
        return False

    head = entry.head(4).lower()
    if any(h in head for h in JOB_TITLE_HINTS):
        return True
    if is_devops_related(text):
        return True

    return False


def extract_experience_entries(pages: Sequence[str]) -> List[Entry]:
    """
    Primary: date-based entries (most robust)
    Fallback: heading-based capture
    """
    date_entries = build_date_based_entries(pages)
    date_entries = [e for e in date_entries if is_experience_entry(e)]
    if date_entries:
        return date_entries

    heading_lines = capture_experience_by_heading(pages)
    if not heading_lines:
        return []

    heading_entries = split_entries_from_lines(heading_lines)
    heading_entries = [e for e in heading_entries if is_experience_entry(e)]
    return heading_entries


# -----------------------------
# Evidence + date math
# -----------------------------

def find_keyword_in_entries(entries: List[Entry], keyword: str) -> Optional[Tuple[int, str]]:
    pattern = re.compile(re.escape(keyword), re.IGNORECASE)
    for entry in entries:
        for idx, (page_num, line) in enumerate(entry.lines):
            if pattern.search(line):
                lines = [l for _, l in entry.lines]
                start = max(idx - SNIPPET_CONTEXT_LINES, 0)
                end = min(idx + SNIPPET_CONTEXT_LINES + 2, len(lines))
                snippet = "\n".join(lines[start:end]).strip()
                return page_num, snippet
    return None


def parse_month_year(token: str, is_start: bool) -> Tuple[Optional[dt.date], bool]:
    token = token.strip().lower()
    today = dt.date.today()

    if token in {"present", "current", "now"}:
        return today, False

    # Year-only: conservative bound
    if re.fullmatch(r"\d{4}", token):
        year = int(token)
        month = 12 if is_start else 1
        return dt.date(year, month, 1), True

    if re.fullmatch(r"\d{1,2}[/-]\d{4}", token):
        month_str, year_str = re.split(r"[/-]", token)
        return dt.date(int(year_str), int(month_str), 1), False

    parts = token.split()
    if len(parts) == 2:
        m, y = parts
        if m in MONTHS and y.isdigit():
            return dt.date(int(y), MONTHS[m], 1), False

    return None, True


def parse_date_ranges(text: str) -> List[Tuple[dt.date, dt.date, bool]]:
    ranges: List[Tuple[dt.date, dt.date, bool]] = []
    for m in DATE_RANGE_PATTERN.finditer(text):
        s_raw, e_raw = m.group("start"), m.group("end")
        s, s_amb = parse_month_year(s_raw, is_start=True)
        e, e_amb = parse_month_year(e_raw, is_start=False)
        if s and e and e >= s:
            ranges.append((s, e, s_amb or e_amb))
    return ranges


def months_between(start: dt.date, end: dt.date) -> List[dt.date]:
    months: List[dt.date] = []
    cur = dt.date(start.year, start.month, 1)
    last = dt.date(end.year, end.month, 1)

    while cur <= last:
        months.append(cur)
        y = cur.year + (cur.month // 12)
        m = cur.month % 12 + 1
        cur = dt.date(y, m, 1)

    return months


def compute_devops_roles(entries: List[Entry]) -> Tuple[List[Role], int, bool]:
    """
    Returns:
    - roles counted (with months_added after overlap removal)
    - total unique DevOps months
    - ambiguity flag (any ambiguous date parsing)
    """
    roles: List[Role] = []
    total_months: Set[dt.date] = set()
    ambiguity = False

    dated: List[Tuple[Entry, dt.date, dt.date, bool]] = []
    for e in entries:
        if not is_devops_related(e.text()):
            continue
        drs = parse_date_ranges(e.text())
        if not drs:
            ambiguity = True
            continue
        for s, en, amb in drs:
            dated.append((e, s, en, amb))

    dated.sort(key=lambda x: x[1])

    for entry, start, end, amb in dated:
        added = 0
        for month in months_between(start, end):
            if month not in total_months:
                total_months.add(month)
                added += 1
        roles.append(Role(title=entry.head(2) or "Unknown title", start=start, end=end, months_added=added))
        ambiguity = ambiguity or amb

    return roles, len(total_months), ambiguity


def months_to_years(months: int) -> float:
    # Display years with 2 decimals (e.g., 2.08 years)
    return round(months / 12.0, 2)


# -----------------------------
# Screening + outputs (DYNAMIC REQUIRED EXPERIENCE)
# -----------------------------

def classify_bucket(result: Dict[str, object]) -> str:
    """
    Bucket is for folder distribution (not Excel):
    - ambiguous: ambiguity == True
    - passed: passed == True and not ambiguous
    - failed: otherwise
    """
    if bool(result.get("ambiguity")):
        return "ambiguous"
    return "passed" if bool(result.get("passed")) else "failed"


def excel_result_label(result: Dict[str, object]) -> str:
    """
    Value shown in the Excel 'result' cell.
    """
    return "AMBIGUOUS" if bool(result.get("ambiguity")) else ("PASS" if bool(result.get("passed")) else "FAIL")


def normalize_excel_col_name(keyword: str) -> str:
    s = keyword.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s or "keyword"


def screen_pdf(pdf_path: Path) -> Dict[str, object]:
    pages, used_ocr = extract_text_by_page(pdf_path)
    exp_entries = extract_experience_entries(pages)

    excluded_entries: List[str] = []
    filtered_entries: List[Entry] = []
    for e in exp_entries:
        if is_excluded(e.text()):
            excluded_entries.append(e.head(3))
        else:
            filtered_entries.append(e)

    required_evidence: Dict[str, Optional[Tuple[int, str]]] = {
        kw: find_keyword_in_entries(filtered_entries, kw) for kw in REQUIRED_EXPERIENCE_KEYWORDS
    }
    all_required_found = all(ev is not None for ev in required_evidence.values())

    roles, devops_months, ambiguity = compute_devops_roles(filtered_entries)
    devops_years = months_to_years(devops_months)

    devops_pass = (devops_years >= MIN_DEVOPS_YEARS) and (not ambiguity)
    passed = all_required_found and devops_pass

    flattened: Dict[str, object] = {}
    for kw in sorted(REQUIRED_EXPERIENCE_KEYWORDS):
        col = normalize_excel_col_name(kw)
        ev = required_evidence.get(kw)
        flattened[f"kw_found__{col}"] = ev is not None
        flattened[f"kw_page__{col}"] = ev[0] if ev else None
        flattened[f"kw_snippet__{col}"] = ev[1] if ev else ""

    # IMPORTANT: always set this as an int (never None)
    experience_entries_found = int(len(filtered_entries))

    return {
        "file": pdf_path.name,
        "passed": passed,
        "required_evidence": required_evidence,
        "devops_years": float(devops_years),
        "devops_roles": roles,
        "excluded_entries": excluded_entries,
        "used_ocr": used_ocr,
        "ambiguity": ambiguity,
        "devops_pass": devops_pass,
        "experience_entries_found": experience_entries_found,
        **flattened,
    }


def build_dynamic_headers() -> List[str]:
    headers: List[str] = ["file", "result"]
    for kw in sorted(REQUIRED_EXPERIENCE_KEYWORDS):
        base = normalize_excel_col_name(kw)
        headers.extend([f"{base}_found", f"{base}_page", f"{base}_snippet"])
    headers.extend(
        [
            "devops_years",
            "devops_pass",
            "date_ambiguity",
            "used_ocr",
        ]
    )
    return headers


def row_for_result(r: Dict[str, object], headers: List[str]) -> List[object]:
    row: List[object] = []
    result_label = excel_result_label(r)

    for h in headers:
        if h == "file":
            row.append(r.get("file"))
        elif h == "result":
            row.append(result_label)
        elif h.endswith("_found") and h not in {"devops_pass"}:
            key = f"kw_found__{h[:-6]}"
            row.append(bool(r.get(key, False)))
        elif h.endswith("_page"):
            key = f"kw_page__{h[:-5]}"
            row.append(r.get(key))
        elif h.endswith("_snippet"):
            key = f"kw_snippet__{h[:-8]}"
            row.append(r.get(key, ""))
        elif h == "devops_years":
            row.append(r.get("devops_years"))
        elif h == "devops_pass":
            row.append(bool(r.get("devops_pass", False)))
        elif h == "date_ambiguity":
            row.append(bool(r.get("ambiguity", False)))
        elif h == "used_ocr":
            row.append(bool(r.get("used_ocr", False)))
        elif h == "experience_entries_found":
            # FIX: never write None; default to 0
            row.append(int(r.get("experience_entries_found", 0)))
        else:
            row.append(r.get(h))
    return row


def write_csv(results: List[Dict[str, object]], output_path: Path) -> None:
    headers = build_dynamic_headers()
    with output_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in results:
            w.writerow(row_for_result(r, headers))


def format_date(d: dt.date) -> str:
    return d.strftime("%Y-%m")


def write_report(results: List[Dict[str, object]], output_path: Path) -> None:
    total = len(results)
    passed = sum(1 for r in results if bool(r["passed"]) and not bool(r.get("ambiguity")))
    ambiguous = sum(1 for r in results if bool(r.get("ambiguity")))
    failed = total - passed - ambiguous

    lines: List[str] = []
    lines.append("# CV Screening Report\n")
    lines.append("## Summary")
    lines.append(f"- Total CVs: {total}")
    lines.append(f"- Passed: {passed}")
    lines.append(f"- Failed: {failed}")
    lines.append(f"- Ambiguous: {ambiguous}\n")

    lines.append("## Active Screening Criteria")
    lines.append(f"- Required keywords in Experience: {', '.join(sorted(REQUIRED_EXPERIENCE_KEYWORDS))}")
    lines.append(f"- Minimum DevOps experience: {MIN_DEVOPS_YEARS} years\n")

    for r in results:
        lines.append(f"## {r['file']}")
        lines.append(f"- Result: {excel_result_label(r)}")
        if r["used_ocr"]:
            lines.append("- Note: OCR fallback used for text extraction.")

        required_evidence: Dict[str, Optional[Tuple[int, str]]] = r["required_evidence"]  # type: ignore
        lines.append("- Required keywords evidence (Experience):")
        for kw in sorted(REQUIRED_EXPERIENCE_KEYWORDS):
            ev = required_evidence.get(kw)
            if ev:
                lines.append(f"  - {kw}: Yes (page {ev[0]})")
                lines.append("    Snippet:\n\n    " + ev[1].replace("\n", "\n    "))
            else:
                lines.append(f"  - {kw}: No")

        lines.append(f"- DevOps years counted (unique, overlap-safe): {r['devops_years']}")
        lines.append(
            f"- DevOps pass (>= {MIN_DEVOPS_YEARS} years AND no ambiguity): {'Yes' if r['devops_pass'] else 'No'}"
        )
        lines.append(f"- Date ambiguity: {'Yes' if r['ambiguity'] else 'No'}")

        roles: List[Role] = r["devops_roles"]  # type: ignore
        if roles:
            lines.append("- DevOps roles counted:")
            for role in roles:
                role_years = months_to_years(role.months_added)
                lines.append(
                    f"  - {role.title} ({format_date(role.start)} to {format_date(role.end)}): {role_years} years"
                )
        else:
            lines.append("- DevOps roles counted: None")

        excl: List[str] = r["excluded_entries"]  # type: ignore
        if excl:
            lines.append("- Excluded entries (ITI/NTI/Sprints/DEPI):")
            for e in excl:
                lines.append(f"  - {e}")
        else:
            lines.append("- Excluded entries (ITI/NTI/Sprints/DEPI): None")

        lines.append("")

    output_path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


# -----------------------------
# Folder distribution (IDEMPOTENT)
# -----------------------------

def ensure_bucket_dirs(output_root: Path) -> Dict[str, Path]:
    passed_dir = output_root / "passed_cvs"
    failed_dir = output_root / "failed_cvs"
    ambiguous_dir = output_root / "ambiguous_cvs"

    passed_dir.mkdir(parents=True, exist_ok=True)
    failed_dir.mkdir(parents=True, exist_ok=True)
    ambiguous_dir.mkdir(parents=True, exist_ok=True)

    return {"passed": passed_dir, "failed": failed_dir, "ambiguous": ambiguous_dir}


def already_distributed(filename: str, bucket_dirs: Dict[str, Path]) -> bool:
    """
    Idempotency rule:
    If the file name exists in ANY bucket folder, skip distribution entirely.
    """
    for d in bucket_dirs.values():
        if (d / filename).exists():
            return True
    return False


def copy_if_absent(src: Path, dst_dir: Path) -> None:
    """
    Copy only if the destination file does not already exist.
    No suffixing is performed.
    """
    dst = dst_dir / src.name
    if dst.exists():
        return
    shutil.copy2(src, dst)


def distribute_pdfs(results: List[Dict[str, object]], input_folder: Path, output_root: Path) -> None:
    bucket_dirs = ensure_bucket_dirs(output_root)

    for r in results:
        filename = str(r["file"])
        pdf_path = input_folder / filename
        if not pdf_path.exists():
            continue

        # Idempotency: if it already exists in any bucket, do nothing (no duplicates).
        if already_distributed(filename, bucket_dirs):
            continue

        bucket = classify_bucket(r)
        copy_if_absent(pdf_path, bucket_dirs[bucket])


# -----------------------------
# Excel output (formatted + colored Result) — DYNAMIC
# -----------------------------

def write_excel(results: List[Dict[str, object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    headers = build_dynamic_headers()
    ws.append(headers)

    for r in results:
        ws.append(row_for_result(r, headers))

    header_font = Font(bold=True)
    top = Alignment(vertical="top")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = top

    ws.freeze_panes = "A2"

    # Explicit filter range (prevents odd extra-column behavior in some cases)
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    header_to_col = {headers[i]: i + 1 for i in range(len(headers))}

    snippet_cols = [h for h in headers if h.endswith("_snippet")]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = top
        for h in snippet_cols:
            col_idx = header_to_col[h]
            row[col_idx - 1].alignment = wrap_top

    result_col = header_to_col["result"]
    fill_pass = PatternFill(fill_type="solid", fgColor="00A000")
    fill_fail = PatternFill(fill_type="solid", fgColor="C00000")
    fill_amb = PatternFill(fill_type="solid", fgColor="F39C12")
    font_white = Font(color="FFFFFF", bold=True)

    for rr in range(2, ws.max_row + 1):
        cell = ws.cell(row=rr, column=result_col)
        val = (cell.value or "").strip().upper()
        if val == "PASS":
            cell.fill = fill_pass
            cell.font = font_white
        elif val == "FAIL":
            cell.fill = fill_fail
            cell.font = font_white
        elif val == "AMBIGUOUS":
            cell.fill = fill_amb
            cell.font = font_white
        cell.alignment = Alignment(horizontal="center", vertical="top")

    # Column widths (auto-ish, with caps)
    caps = {h: 45 for h in headers}
    caps["file"] = 55
    caps["result"] = 14
    caps["devops_years"] = 14
    for h in snippet_cols:
        caps[h] = 80

    for col_idx, h in enumerate(headers, start=1):
        longest = len(h)
        for rr in range(2, ws.max_row + 1):
            v = ws.cell(row=rr, column=col_idx).value
            if v is None:
                continue
            longest = max(longest, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(caps.get(h, 45), max(10, longest + 2))

    snippet_col_indices = [header_to_col[h] for h in snippet_cols]
    for rr in range(2, ws.max_row + 1):
        long_snip = False
        for col in snippet_col_indices:
            val = ws.cell(row=rr, column=col).value or ""
            if len(str(val)) > 80:
                long_snip = True
                break
        ws.row_dimensions[rr].height = 70 if long_snip else 20

    wb.save(output_path)

def apply_cli_overrides(min_devops_years: Optional[float], required_keywords: Optional[List[str]]) -> None:
    """
    Apply CLI overrides to the global EASY CONFIG values.
    """
    global MIN_DEVOPS_YEARS, REQUIRED_EXPERIENCE_KEYWORDS

    if min_devops_years is not None:
        MIN_DEVOPS_YEARS = float(min_devops_years)

    if required_keywords is not None and len(required_keywords) > 0:
        REQUIRED_EXPERIENCE_KEYWORDS = set(required_keywords)



# -----------------------------
# Main
# -----------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="Screen CV PDFs for DevOps requirements.")
    parser.add_argument("folder", nargs="?", default="./cvs", help="Folder containing PDF CVs (default: ./cvs)")
    parser.add_argument("--output-dir", default=".", help="Output directory (default: current directory)")

    # CLI overrides for EASY CONFIG
    parser.add_argument(
        "--min-devops-years",
        type=float,
        default=None,
        help="Override MIN_DEVOPS_YEARS (e.g., 3.0)",
    )
    parser.add_argument(
        "--required-keyword",
        action="append",
        default=None,
        help="Add a required experience keyword (repeatable). Example: --required-keyword Kubernetes --required-keyword AWS",
    )

    args = parser.parse_args()

    # Apply overrides (if any)
    apply_cli_overrides(args.min_devops_years, args.required_keyword)

    folder = Path(args.folder).resolve()
    outdir = Path(args.output_dir).resolve()
    outdir.mkdir(parents=True, exist_ok=True)

    pdfs = sorted(folder.glob("*.pdf"))
    results: List[Dict[str, object]] = [screen_pdf(pdf) for pdf in pdfs]

    # Output files
    write_csv(results, outdir / "screening_results.csv")
    write_report(results, outdir / "screening_report.md")
    write_excel(results, outdir / "screening_results.xlsx")

    # Folder distribution (idempotent)
    distribute_pdfs(results, folder, outdir)


if __name__ == "__main__":
    main()
